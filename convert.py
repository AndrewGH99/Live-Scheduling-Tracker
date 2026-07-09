#!/usr/bin/env python3
"""
RPDC Schedule Converter  —  tour + station-priority edition
===========================================================
Reads a simple roster spreadsheet (INPUT_TEMPLATE.xlsx) — where each FTC
confirms the hours they can work — and produces the polished GH Logistics
Memphis RPDC weekly board (RPDC_Board.xlsx):

  - Station + Contact/Badge columns carried through
  - Auto-computed weekly hours (overnight-aware, split-shift-aware)
  - COVERAGE scored by USPS tour against the SOP manpower-priority model:
        Tour 1  = 11:00 PM – 6:00 AM   (overnight)
        Tour 2  =  6:00 AM – 2:00 PM   (day)
        Tour 3  =  2:00 PM – 10:00 PM  (evening)
  - Per SOP "Staffing Priority 24/7/365", bodies-on-tour -> stations manned:
        4+ -> #1 Floor Lead · #2 Docks · #3 Yard/Inbound · #4 Overflow  (green)
        3  -> #1 · #2 · #3        (Overflow down)                       (yellow)
        2  -> #1 · #2             (Yard + Overflow down)                (orange)
        1  -> #1 only             (critical)                            (red)
        0  -> NONE — breaches the 24/7 Station #1 mandate               (black)
  - GAPS block: every tour x day below the 4-body baseline, worst first.

Usage:
    python3 convert.py                       # INPUT_TEMPLATE.xlsx -> RPDC_Board.xlsx
    python3 convert.py roster.xlsx           # roster.xlsx        -> RPDC_Board.xlsx
    python3 convert.py roster.xlsx out.xlsx  # explicit in and out

The roster->board logic is isolated here so a future live version (shared
sheet / Hubstaff / Zoom presence) swaps only the *input* reader.
"""
import sys
import re
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
BASELINE = 4  # bodies/tour target = all 4 stations manned

# USPS tours (start, end) on a 24h clock; end<=start means it wraps midnight
TOURS = [
    ("Tour 1", "11p–6a",  (23.0, 6.0)),
    ("Tour 2", "6a–2p",   (6.0, 14.0)),
    ("Tour 3", "2p–10p",  (14.0, 22.0)),
]
COVER_FRACTION = 0.5  # a body "covers" a tour if on-clock >= half the tour

STATION_NAMES = ["#1 Floor Lead", "#2 Docks", "#3 Yard/Inbound", "#4 Overflow"]

# ---- shared definitions (SOP) — reused in the board's Legend sheet ----------
STATION_DEFS = [
    ("#1 · Inside/Outside Floor Lead",
     "ALWAYS staffed 24/7/365. Runs the interior floor (primary) + exterior docks "
     "(secondary), first responder inside & out, owns driver badging / PPE / check-in. "
     "The one station that can never be empty."),
    ("#2 · Outside Docks",
     "Always staffed. Continuous dock patrol, hourly chock + trailer audits, meets every "
     "arriving truck, collects straps, enforces driver compliance (padlock / vest / closed-toe "
     "= fines), ensures trucks are unhooked & sealed before departure."),
    ("#3 · Trailer Yard / Inbound",
     "Owns inbound trailer inspection (four-photo process), running yard inventory / audit, "
     "flags out-of-service units, daily technician check, moves trailers."),
    ("#4 · Overflow / Support",
     "Activated only when a 4th body is available. Backs up #2 & #3, carrier hunting (inbound "
     "engagement), trailer decaling, owns the 3 PM / 3 AM reactivated-trailer list."),
]
# (color, bodies, what it means / why) — from SOP "Staffing Priority 24/7/365"
COLOR_DEFS = [
    ("Green",  "4+", "All 4 stations manned — full operation with Overflow/surge capacity."),
    ("Yellow", "3",  "#1 · #2 · #3 covered; Overflow (#4) down. Core ops fine, no surge capacity."),
    ("Orange", "2",  "#1 · #2 only. Trailer Yard/Inbound (#3) unmanned — inbound four-photo "
                     "inspection at risk."),
    ("Red",    "1",  "Floor Lead (#1) only; docks unmanned. Critical single point of failure."),
    ("Black",  "0",  "No one on the tour — breaches the 24/7 Station #1 mandate."),
]
TOUR_DEFS = [
    ("Tour 1", "11:00 PM – 6:00 AM", "overnight"),
    ("Tour 2", "6:00 AM – 2:00 PM",  "day"),
    ("Tour 3", "2:00 PM – 10:00 PM", "evening"),
]

# Cells that mean "not working": blank / off / x / -
OFF_TOKENS = {"", "off", "x", "-", "n", "no", "none"}
# Recognises 'Tour 2', 'T3', 'tour1' entries so the sheet can be filled by tour name
TOUR_LABEL_RE = re.compile(r"\bt(?:our)?\s*([123])\b", re.I)

# ---- styling ---------------------------------------------------------------
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_ORANGE = PatternFill("solid", fgColor="FFCC99")
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")
FILL_BLACK  = PatternFill("solid", fgColor="808080")
FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FILL_TITLE  = PatternFill("solid", fgColor="2E75B6")
FILL_SCHED  = PatternFill("solid", fgColor="DDEBF7")
FILL_COVLBL = PatternFill("solid", fgColor="D9D9D9")
FILL_GAPHDR = PatternFill("solid", fgColor="C00000")

WHITE_BOLD  = Font(bold=True, color="FFFFFF")
BOLD        = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---- clock / hour parsing --------------------------------------------------
def _parse_clock(tok, inherited_mer):
    tok = tok.strip().lower().replace(" ", "")
    mer = None
    if tok.endswith("a"):
        mer, tok = "a", tok[:-1]
    elif tok.endswith("p"):
        mer, tok = "p", tok[:-1]
    if ":" in tok:
        h, m = tok.split(":")
        val = int(h) + int(m) / 60.0
    else:
        val = float(int(tok))
    use = mer or inherited_mer
    hour12 = val % 12
    if use == "p":
        hour24 = hour12 + 12
    elif use == "a":
        hour24 = hour12
    else:
        hour24 = val
    return hour24, mer


def _segment_range(seg):
    """'9a–6p' -> (9.0, 18.0). Overnight kept as-is (end<=start)."""
    parts = [p for p in re.split(r"[–—\-]", seg) if p.strip()]
    if len(parts) != 2:
        return None
    _, end_mer = _parse_clock(parts[1], None)
    start_h, _ = _parse_clock(parts[0], end_mer)
    end_h, _ = _parse_clock(parts[1], None)
    return (start_h, end_h)


def _segment_hours(seg):
    r = _segment_range(seg)
    if not r:
        return 0.0
    dur = r[1] - r[0]
    if dur <= 0:
        dur += 24
    return dur


def _norm(val):
    return str(val).strip().lower() if val is not None else ""


def is_scheduled(val):
    return _norm(val) not in OFF_TOKENS


def explicit_tours(val):
    """0-based tour indices if the cell names tours ('Tour 2', 'T1'); [] if
    explicitly off/blank; None if the cell is hours or other free text."""
    s = _norm(val)
    if s in OFF_TOKENS:
        return []
    found = TOUR_LABEL_RE.findall(s)
    if found:
        return sorted({int(x) - 1 for x in found})
    return None


def cell_hours(val):
    et = explicit_tours(val)
    if et is not None:
        return round(sum(_tour_len(TOURS[i][2]) for i in et), 2)
    s = str(val).strip() if val is not None else ""
    if not s or not re.search(r"\d", s) or not re.search(r"[–—\-]", s):
        return 0.0
    return round(sum(_segment_hours(seg) for seg in s.split("/")), 2)


# ---- tour overlap ----------------------------------------------------------
def _ivs(s, e):
    """Half-open interval(s) for [s,e); splits across midnight if e<=s."""
    return [(s, e)] if e > s else [(s, 24.0), (0.0, e)]


def _tour_len(win):
    return sum(hi - lo for lo, hi in _ivs(*win))


def tour_overlap(cell, win):
    """Hours a schedule cell overlaps a tour window. 'Avail.'/non-time = full."""
    if not is_scheduled(cell):
        return 0.0
    s = str(cell).strip()
    tws = _ivs(*win)
    if not re.search(r"[–—\-]", s):        # 'Avail.' -> always-on layer
        return _tour_len(win)
    total = 0.0
    for seg in s.split("/"):
        rng = _segment_range(seg)
        if not rng:
            continue
        for pv in _ivs(rng[0], rng[1]):
            for tw in tws:
                total += max(0.0, min(pv[1], tw[1]) - max(pv[0], tw[0]))
    return total


def covers_tour(cell, ti):
    """Does this cell cover tour index ti? Honours 'Tour N' labels, else hours."""
    et = explicit_tours(cell)
    if et is not None:
        return ti in et
    win = TOURS[ti][2]
    return tour_overlap(cell, win) >= COVER_FRACTION * _tour_len(win)


def tour_count(rows, day_idx, ti):
    return sum(1 for p in rows if covers_tour(p["days"][day_idx], ti))


def cov_fill(n):
    return (FILL_GREEN if n >= 4 else FILL_YELLOW if n == 3
            else FILL_ORANGE if n == 2 else FILL_RED if n == 1 else FILL_BLACK)


def stations_label(n):
    if n >= 4:
        return "All 4 stations"
    if n == 3:
        return "#1·#2·#3 (Overflow down)"
    if n == 2:
        return "#1·#2 (Yard+Overflow down)"
    if n == 1:
        return "#1 only (CRITICAL)"
    return "NONE — 24/7 breach"


# ---- input reader (swap this for a live source) ----------------------------
def read_roster(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Roster"] if "Roster" in wb.sheetnames else wb.worksheets[0]
    header_row = None
    for r in range(1, ws.max_row + 1):
        vals = [str(c.value).strip().lower() if c.value else "" for c in ws[r]]
        if "employee name" in vals:
            header_row = r
            break
    if header_row is None:
        raise SystemExit("Could not find an 'Employee Name' header row in the roster.")
    headers = [str(c.value).strip() if c.value else "" for c in ws[header_row]]
    idx = {h.lower(): i for i, h in enumerate(headers)}

    def col(row, *names):
        for n in names:
            if n in idx and idx[n] < len(row):
                return row[idx[n]]
        return None

    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = [c.value for c in ws[r]]
        name = col(row, "employee name")
        if not name or not str(name).strip():
            continue
        rows.append({
            "name":    str(name).strip(),
            "contact": col(row, "contact/badge", "contact", "badge"),
            "role":    col(row, "role") or "Facilities Transportation Coordinator",
            "shift":   col(row, "shift") or "",
            "station": col(row, "station") or "",
            "days":    [col(row, d.lower()) for d in DAYS],
            "notes":   col(row, "notes") or "",
        })
    return rows


# ---- board writer ----------------------------------------------------------
def build_board(rows, out_path, updated=None):
    updated = updated or datetime.date.today().strftime("%-m/%-d/%Y")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Memphis RPDC Schedule"

    cols = ["Employee Name", "Contact/Badge", "Role", "Shift", "Station"] + DAYS + ["Notes", "Wk Hrs"]
    ncol = len(cols)
    last_col = get_column_letter(ncol)
    day_start = 6

    # Title
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "GH Logistics — Memphis RPDC · Onsite Weekly Schedule & Tour Coverage"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = FILL_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Subtitle
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = (f"Updated {updated}. Coverage scored per SOP manpower-priority (Stations #1–4). "
               f"Tours: T1 11p–6a · T2 6a–2p · T3 2p–10p. Baseline = {BASELINE} bodies/tour "
               f"(all 4 stations). A body counts toward a tour if on-clock ≥ half of it. "
               f"'Avail.' = 24/7 always-on layer, counts in every tour.")
    c.font = Font(italic=True, size=9)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 42

    # Header
    hdr = 3
    for j, name in enumerate(cols, start=1):
        cc = ws.cell(row=hdr, column=j, value=name)
        cc.font = WHITE_BOLD
        cc.fill = FILL_HEADER
        cc.alignment = CENTER
        cc.border = BORDER

    # People
    r = hdr + 1
    for row in rows:
        wk = sum(cell_hours(d) for d in row["days"])
        vals = [row["name"], row["contact"], row["role"], row["shift"], row["station"]] \
            + row["days"] + [row["notes"], round(wk, 2) if wk else 0]
        for j, v in enumerate(vals, start=1):
            cc = ws.cell(row=r, column=j, value=v)
            cc.border = BORDER
            if j == 3 or j == day_start + 7:
                cc.alignment = LEFT
            else:
                cc.alignment = CENTER
            if day_start <= j <= day_start + 6 and is_scheduled(v):
                cc.fill = FILL_SCHED
        r += 1

    # Coverage by tour
    r += 1
    ct = ws.cell(row=r, column=1, value="TOUR COVERAGE (bodies on-tour → stations manned)")
    ct.font = BOLD
    r += 1
    for ti, (tname, twin_lbl, win) in enumerate(TOURS):
        lc = ws.cell(row=r, column=4, value=f"{tname} {twin_lbl} →")
        lc.font = BOLD
        lc.alignment = Alignment(horizontal="right", vertical="center")
        for k in (1, 2, 3, 5):
            ws.cell(row=r, column=k).fill = FILL_COVLBL
        lc.fill = FILL_COVLBL
        for di in range(7):
            n = tour_count(rows, di, ti)
            cc = ws.cell(row=r, column=day_start + di, value=n)
            cc.alignment = CENTER
            cc.font = Font(bold=True, color="FFFFFF" if n == 0 else "000000")
            cc.border = BORDER
            cc.fill = cov_fill(n)
        r += 1

    # Gaps summary
    gaps = []
    for ti, (tname, twin_lbl, win) in enumerate(TOURS):
        for di in range(7):
            n = tour_count(rows, di, ti)
            if n < BASELINE:
                gaps.append((n, tname, twin_lbl, DAYS[di], BASELINE - n))
    gaps.sort(key=lambda g: (g[0], DAYS.index(g[3])))

    r += 1
    gh = ws.cell(row=r, column=1, value=f"GAPS — where you are short of {BASELINE} (worst first)")
    gh.font = Font(bold=True, color="FFFFFF")
    gh.fill = FILL_GAPHDR
    for k in range(2, 7):
        ws.cell(row=r, column=k).fill = FILL_GAPHDR
    r += 1
    if not gaps:
        ws.cell(row=r, column=1, value="None — every tour × day meets the 4-body baseline. 🟢").font = BOLD
        r += 1
    else:
        for j, h in enumerate(["Tour", "Window", "Day", "Bodies", "Short by", "Stations manned"], start=1):
            hc = ws.cell(row=r, column=j, value=h)
            hc.font = WHITE_BOLD
            hc.fill = FILL_HEADER
            hc.border = BORDER
        r += 1
        for n, tname, twin_lbl, day, short in gaps:
            vals = [tname, twin_lbl, day, n, short, stations_label(n)]
            for j, v in enumerate(vals, start=1):
                cc = ws.cell(row=r, column=j, value=v)
                cc.border = BORDER
                cc.alignment = LEFT if j == 6 else CENTER
                if j == 4:
                    cc.fill = cov_fill(n)
                    cc.font = Font(bold=True, color="FFFFFF" if n == 0 else "000000")
            r += 1

    # Legend
    r += 1
    ws.cell(row=r, column=1, value=(
        "Legend:  ≥4 green = all 4 stations · 3 yellow = #1·#2·#3 · 2 orange = #1·#2 · "
        "1 red = #1 only (critical) · 0 black = 24/7 breach")).font = Font(italic=True, size=9)

    # Widths
    widths = {1: 18, 2: 13, 3: 34, 4: 14, 5: 24, 13: 40, 14: 8}
    for di in range(7):
        widths[day_start + di] = 14
    for j, w in widths.items():
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "F4"
    build_legend(wb)
    wb.save(out_path)
    return out_path, gaps


def build_legend(wb):
    """A plain-English 'Legend & Definitions' sheet: stations, colors, tours."""
    ls = wb.create_sheet("Legend & Definitions")
    color_fill = {"Green": FILL_GREEN, "Yellow": FILL_YELLOW, "Orange": FILL_ORANGE,
                  "Red": FILL_RED, "Black": FILL_BLACK}
    r = 1

    def section(title):
        nonlocal r
        c = ls.cell(row=r, column=1, value=title)
        c.font = Font(bold=True, size=13, color="FFFFFF")
        c.fill = FILL_TITLE
        ls.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    section("THE 4 STATIONS  (SOP — filled in this priority order)")
    for name, desc in STATION_DEFS:
        nc = ls.cell(row=r, column=1, value=name)
        nc.font = BOLD
        nc.alignment = Alignment(vertical="top", wrap_text=True)
        dc = ls.cell(row=r, column=2, value=desc)
        dc.alignment = Alignment(wrap_text=True, vertical="top")
        ls.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        ls.row_dimensions[r].height = 46
        r += 1

    r += 1
    section("WHY RED / YELLOW / ORANGE / GREEN  (bodies on a tour → stations manned)")
    hdr = ["Color", "Bodies", "What it means"]
    for j, h in enumerate(hdr, start=1):
        hc = ls.cell(row=r, column=j, value=h)
        hc.font = WHITE_BOLD
        hc.fill = FILL_HEADER
    r += 1
    for color, bodies, meaning in COLOR_DEFS:
        cc = ls.cell(row=r, column=1, value=color)
        cc.fill = color_fill[color]
        cc.font = Font(bold=True, color="FFFFFF" if color == "Black" else "000000")
        ls.cell(row=r, column=2, value=bodies).alignment = CENTER
        mc = ls.cell(row=r, column=3, value=meaning)
        mc.alignment = Alignment(wrap_text=True, vertical="top")
        ls.row_dimensions[r].height = 30
        r += 1

    r += 1
    section("THE 3 TOURS")
    for tname, window, kind in TOUR_DEFS:
        ls.cell(row=r, column=1, value=tname).font = BOLD
        ls.cell(row=r, column=2, value=window)
        ls.cell(row=r, column=3, value=kind)
        r += 1

    r += 1
    ls.cell(row=r, column=1, value=(
        "Baseline = 4 bodies per tour (all 4 stations). A body counts toward a tour if "
        "on-clock for at least half of it. 'Avail.' = the 24/7 lead, counts in every tour.")
    ).alignment = Alignment(wrap_text=True)
    ls.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    ls.column_dimensions["A"].width = 30
    ls.column_dimensions["B"].width = 12
    ls.column_dimensions["C"].width = 70


def main():
    inp = sys.argv[1] if len(sys.argv) > 1 else "INPUT_TEMPLATE.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "RPDC_Board.xlsx"
    rows = read_roster(inp)
    path, gaps = build_board(rows, out)
    print(f"✓ Built {path} from {inp}  ({len(rows)} people)")
    if gaps:
        print(f"  {len(gaps)} tour×day gap(s) below baseline {BASELINE}:")
        for n, t, w, d, short in gaps:
            print(f"    {t} {w}  {d}: {n} bodies (short {short}) — {stations_label(n)}")
    else:
        print("  No coverage gaps — all tours meet baseline.")


if __name__ == "__main__":
    main()
