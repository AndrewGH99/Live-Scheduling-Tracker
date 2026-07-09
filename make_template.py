#!/usr/bin/env python3
"""
Generates INPUT_TEMPLATE.xlsx — the dead-simple roster the RPDC team fills to
confirm the hours they can work. Every day cell is a CLICK-a-dropdown (Off /
Tour 1 / Tour 2 / Tour 3 / Avail.), so nobody has to type. Custom exact hours
are still allowed. Includes an INSTRUCTIONS tab and a shared 'Legend &
Definitions' tab (stations + why Red/Yellow/Orange/Green), reused verbatim from
convert.py so the two never drift.

Run:  python3 make_template.py            -> INPUT_TEMPLATE.xlsx  (RPDC)
      python3 make_template.py LC         -> INPUT_TEMPLATE_LC.xlsx (Load Coordinators, blank)
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import convert  # reuse the Legend/definitions so they stay in sync

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
COLS = ["Employee Name", "Contact/Badge", "Role", "Shift", "Station"] + DAYS + ["Notes"]

# dropdown option lists (values chosen so convert.py parses them directly)
DAY_OPTS     = "Off,Tour 1 (11p–6a),Tour 2 (6a–2p),Tour 3 (2p–10p),Avail."
SHIFT_OPTS   = "Day,Swing,Night,Day & Night"
STATION_OPTS = "Inside/Outside Floor Lead,Outside Docks,Trailer Yard/Inbound,Overflow"

FTC = "Facilities Transportation Coordinator"
# Fictional demo crew — same shift patterns as a real 11-person site, so the
# coverage math is realistic. Load your real roster via Excel or Google Sheet.
CREW = [
    ("Jordan Rivers", "", "Onsite Lead (24/7 Day & Night)", "Day & Night", "Inside/Outside Floor Lead",
     ["Avail."] * 7, "Available daily; flexes to cover weekends."),
    ("Casey Brooks", "", FTC, "Day", "Outside Docks",
     ["9a–6p", "9a–6p", "9a–6p", "9a–6p", "9a–6p", "", ""], "Day-shift supervisor per SOP."),
    ("Devon Hayes", "", FTC, "Day", "Trailer Yard/Inbound",
     ["", "7a–5p", "7a–5p", "7a–5p", "7a–5p", "7a–5p", ""], "Tue–Sat 7 AM–5 PM."),
    ("Morgan Cole", "", FTC, "Day", "Outside Docks",
     ["", "", "6a–4p", "6a–4p", "6a–4p", "6a–4p", "6a–4p"], "Day shift, Wed–Sun."),
    ("Taylor Brant", "", FTC, "Day", "Overflow",
     ["8a–5p", "8a–5p", "8a–5p", "8a–5p", "", "", "8a–5p"], "New hire (starts Mon)."),
    ("Skyler Dunn", "", FTC, "Swing", "Inside/Outside Floor Lead",
     ["2p–11p", "2p–11p", "2p–11p", "2p–11p", "", "", "2p–11p"], "Swing 2 PM–11 PM."),
    ("Riley Vaughn", "", FTC, "Night", "Trailer Yard/Inbound",
     ["10p–8a", "10p–8a", "10p–8a", "10p–8a", "10p–8a", "10p–8a", ""], "Mon–Sat 10 PM–8 AM."),
    ("Kai Emerson", "", FTC, "Night", "Outside Docks",
     ["10–11p / 8–9a"] * 6 + [""], "6-day. NOTE: this split covers no full tour — confirm real hours."),
    ("Blake Connor", "", FTC, "Night", "Inside/Outside Floor Lead",
     ["", "", "10–11p / 8–9a", "10–11p / 8–9a", "10–11p / 8–9a", "10–11p / 8–9a", "10–11p / 8–9a"],
     "Wed–Sun. NOTE: split covers no full tour — confirm real hours."),
    ("Marley West", "", FTC, "Night", "Overflow",
     ["", "", "6p–4a", "6p–4a", "6p–4a", "6p–4a", "6p–4a"], "6 PM–4 AM."),
    ("Avery Nash", "", FTC, "Night", "Trailer Yard/Inbound",
     ["10:30p–7a", "10:30p–7a", "10:30p–7a", "10:30p–7a", "10:30p–7a", "", ""], "BOL / clerk focus."),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MAXROW = 60  # rows given dropdowns / room to grow


def _dv(formula, cells, ws):
    dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
    dv.showErrorMessage = False  # advisory dropdown — custom hours still allowed
    ws.add_data_validation(dv)
    dv.add(cells)


def build(path, crew, role_label):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Roster"

    for j, name in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=j, value=name)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for i, person in enumerate(crew, start=2):
        name, contact, role, shift, station, days, notes = person
        vals = [name, contact, role, shift, station] + days + [notes]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if j in (1, 3, len(COLS)) else "center",
                                    vertical="center", wrap_text=True)

    # dropdowns
    _dv(DAY_OPTS, f"F2:L{MAXROW}", ws)          # Mon..Sun
    _dv(SHIFT_OPTS, f"D2:D{MAXROW}", ws)        # Shift
    _dv(STATION_OPTS, f"E2:E{MAXROW}", ws)      # Station

    widths = {1: 18, 2: 14, 3: 34, 4: 13, 5: 24, 13: 40}
    for di in range(7):
        widths[6 + di] = 16
    for j, w in widths.items():
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

    # INSTRUCTIONS
    ins = wb.create_sheet("INSTRUCTIONS")
    lines = [
        (f"How to fill this out — {role_label}", True),
        ("", False),
        ("One row per person. For each day, CLICK the cell and pick from the dropdown:", False),
        ("   Off  ·  Tour 1 (11p–6a)  ·  Tour 2 (6a–2p)  ·  Tour 3 (2p–10p)  ·  Avail.", False),
        ("That's it — no typing needed to confirm coverage.", False),
        ("", False),
        ("Prefer exact hours? You can still type them instead:  9a–6p · 10p–8a · 6:30a–3p", False),
        ("   Split shift: separate with a slash  ->  10–11p / 8–9a", False),
        ("   Leave a day blank (or pick Off) if the person is off that day.", False),
        ("", False),
        ("Shift & Station columns are dropdowns too. Then save the file and either:", False),
        ("   • drag it into rpdc_live.html, or", False),
        ("   • run:  python3 convert.py INPUT_TEMPLATE.xlsx RPDC_Board.xlsx", False),
        ("", False),
        ("Weekly hours and color-coded tour coverage are computed for you.", False),
        ("See the 'Legend & Definitions' tab for the stations and what the colors mean.", False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = ins.cell(row=i, column=1, value=text)
        c.font = Font(bold=True, size=13) if (bold and i == 1) else Font(bold=bold)
    ins.column_dimensions["A"].width = 92

    # shared Legend & Definitions (identical to the board's)
    convert.build_legend(wb)

    wb.save(path)
    print(f"✓ Wrote {path}  ({len(crew)} rows) — day cells are dropdowns; Legend tab included")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1].upper() == "LC":
        # Load Coordinators: blank starter, same structure (fill in real LC crew later)
        build("INPUT_TEMPLATE_LC.xlsx", [], "Load Coordinators (LC)")
    else:
        build("INPUT_TEMPLATE.xlsx", CREW, "RPDC Facilities Transportation Coordinators (FTC)")
